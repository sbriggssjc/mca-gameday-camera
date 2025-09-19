#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys, json, csv, glob, re
from pathlib import Path
from collections import Counter
from openpyxl import Workbook

def get_out_path():
    out = os.environ.get("OUT") or (sys.argv[1] if len(sys.argv) > 1 else None)
    if not out:
        print("Provide OUT via env or as first CLI arg", file=sys.stderr); sys.exit(2)
    return Path(out)

def read_jsonl(p: Path):
    rows = []
    with p.open(encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s: continue
            try: rows.append(json.loads(s))
            except Exception: pass
    return rows

def read_csv_dict(p: Path):
    if not p.exists(): return []
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def norm_side(v):
    s = str(v or "").strip().lower()
    if s in {"o","off","offense","offence"}: return "offense"
    if s in {"d","def","defense","defence"}: return "defense"
    return s

def distance_bucket(dist):
    try: d = float(dist)
    except Exception: return "unknown"
    if d <= 2: return "0-2"
    if d <= 5: return "3-5"
    if d <= 8: return "6-8"
    if d <= 12: return "9-12"
    return "13+"

def yard_bucket(y):
    try: v = float(y)
    except Exception: return "unknown"
    if v <= -3: return "<= -3"
    if v <= 0:  return "-2 to 0"
    if v <= 3:  return "1 to 3"
    if v <= 7:  return "4 to 7"
    return "8+"

def best_clip_for_index(out_dir: Path, idx):
    cdir = out_dir / "clips"
    if not cdir.exists(): return ""
    idx = str(idx)
    patterns = [f"*_{idx}_*.mp4", f"*_{idx}.mp4", f"{idx}_*.mp4", f"{idx}.mp4", f"*{idx}*.mp4"]
    for pat in patterns:
        files = sorted(glob.glob(str(cdir / pat)))
        if files: return Path(files[0]).name
    return ""

def build_workbook(catalog_rows, offense_rows, defense_rows, xlsx_path: Path):
    wb = Workbook()
    ws_cat = wb.active; ws_cat.title = "catalog"
    front = ["index","clip_file","jenks_side","metro_side",
             "play_type_use","dir_use","down_use","distance_use","gained_yards_use",
             "distance_bucket","yards_bucket","outcome_use"]
    seen, all_keys = set(), []
    for r in catalog_rows:
        for k in r.keys():
            if k not in seen: seen.add(k); all_keys.append(k)
    cols = [c for c in front if c in seen] + [k for k in all_keys if k not in front]
    ws_cat.append(cols)
    for r in catalog_rows: ws_cat.append([r.get(c, "") for c in cols])

    def write_summary_sheet(title, rows):
        ws = wb.create_sheet(title)
        def vc(col_key, label):
            c = Counter()
            for r in rows:
                c[str(r.get(col_key, "") or "")] += 1
            out = [(label, col_key, "count")]
            out += [(label, k, n) for k,n in c.most_common()]
            return out
        blocks = [
            vc("play_type_use","by_playtype"),
            vc("dir_use","by_direction"),
            vc("down_use","by_down"),
            vc("distance_bucket","by_distance_bucket"),
            vc("outcome_use","by_outcome"),
            vc("yards_bucket","by_yards_bucket"),
        ]
        # by play name if present
        play_name_key = None
        for cand in ("play_name","call","play","concept","tag","name"):
            if any(cand in r for r in rows): play_name_key = cand; break
        if play_name_key: blocks.append(vc(play_name_key, "by_play"))
        r0 = 1
        for block in blocks:
            ws.cell(row=r0, column=1, value="summary")
            ws.cell(row=r0, column=2, value=block[0][1])
            ws.cell(row=r0, column=3, value="count"); r0 += 1
            for _, key, n in block[1:]:
                ws.cell(row=r0, column=1, value=block[0][0])
                ws.cell(row=r0, column=2, value=key)
                ws.cell(row=r0, column=3, value=n); r0 += 1
            r0 += 2

    write_summary_sheet("summary_offense", offense_rows)
    write_summary_sheet("summary_defense", defense_rows)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True); wb.save(xlsx_path)

def main():
    out = get_out_path()
    plays_path = out/"plays.jsonl"
    audit_path = out/"audit"/"audit_template.csv"
    export_dir = out/"export"; export_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = export_dir/"opponent_all_plays.xlsx"

    plays = read_jsonl(plays_path)
    audit = read_csv_dict(audit_path)

    # build audit index map
    aud_idx_key = None
    if audit:
        for k in audit[0].keys():
            if k.lower()=="index": aud_idx_key=k; break
        if not aud_idx_key:
            aud_idx_key="index"
            for i,r in enumerate(audit): r["index"]=str(i)
    aud_by_idx = {str(r.get(aud_idx_key,"")).strip(): r for r in audit} if audit else {}

    catalog, offense, defense = [], [], []
    for i, p in enumerate(plays):
        row = dict(p)
        idx = None
        for k in ("index","idx","row_index","row_id","play_index","play_id"):
            if k in row: idx = row[k]; break
        if idx is None: idx = i
        idx_s = str(idx)

        aud = aud_by_idx.get(idx_s, {})
        for k,v in aud.items():
            if k not in row or str(row[k]).strip()=="": row[k]=v

        js = row.get("jenks_side") or norm_side(row.get("side_fix")) or norm_side(row.get("side_auto"))
        ms = "defense" if js=="offense" else ("offense" if js=="defense" else "")
        row["jenks_side"]=js or ""; row["metro_side"]=ms

        clip = row.get("clip") or ""
        if not str(clip).strip() or clip=="null": clip = best_clip_for_index(out, idx_s)
        row["clip_file"]=clip

        rp = (row.get("rp_fix") or row.get("rp_auto") or row.get("play_type") or row.get("type") or "")
        rp = {"r":"run","p":"pass","run/pass":"runpass"}.get(str(rp).lower(), str(rp).lower())
        row["play_type_use"]=rp

        row["dir_use"]= row.get("dir_fix") or row.get("dir_auto") or row.get("direction") or ""
        m = re.search(r"(\d+)", str(row.get("down_fix") or row.get("down") or ""))
        row["down_use"]= m.group(1) if m else ""
        row["distance_use"]= row.get("distance_fix") or row.get("distance") or ""
        row["gained_yards_use"]= row.get("gained_yards_fix") or row.get("gained_yards") or row.get("yards_gained") or row.get("yards") or ""
        row["distance_bucket"]= distance_bucket(row["distance_use"])
        row["yards_bucket"]= yard_bucket(row["gained_yards_use"])

        outcome = row.get("auto_outcome")
        if outcome and str(outcome).strip():
            row["outcome_use"]= outcome
        else:
            try: yg=float(row["gained_yards_use"])
            except Exception: row["outcome_use"]="unknown"
            else:
                row["outcome_use"]= "negative" if yg<=-1 else "short" if yg<=2 else "medium" if yg<=6 else "explosive"

        row["index"]=idx_s
        catalog.append(row)
        if js=="offense": offense.append(row)
        elif js=="defense": defense.append(row)

    build_workbook(catalog, offense, defense, xlsx_path)
    print(f"[export] wrote: {xlsx_path}")

if __name__ == "__main__":
    main()
